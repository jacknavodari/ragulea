import sys
import os
import traceback
from fastapi.responses import FileResponse

# Setup logging to file immediately to catch import errors
app_data_dir = os.path.join(os.getenv('APPDATA'), 'RAGulea')
os.makedirs(app_data_dir, exist_ok=True)
log_file = os.path.join(app_data_dir, "ragulea_debug.log")

def log_error(msg):
    try:
        with open(log_file, "a") as f:
            f.write(msg + "\n")
    except:
        pass

try:
    from fastapi import FastAPI, UploadFile, File, HTTPException, Body
    from fastapi.middleware.cors import CORSMiddleware
    from fastapi.staticfiles import StaticFiles
    from pydantic import BaseModel
    from typing import List, Optional
    import shutil
    import uvicorn
    import socket
    from pymongo import MongoClient
    from langchain_text_splitters import RecursiveCharacterTextSplitter
    from langchain_ollama import OllamaEmbeddings, OllamaLLM
    import numpy as np
    from bson.objectid import ObjectId
    import fitz  # PyMuPDF
    try:
        from docx import Document as DocxDocument
        DOCX_AVAILABLE = True
    except ImportError:
        DOCX_AVAILABLE = False
    try:
        from openpyxl import load_workbook
        EXCEL_AVAILABLE = True
    except ImportError:
        EXCEL_AVAILABLE = False
    try:
        import pytesseract
        from pdf2image import convert_from_path
        from PIL import Image
        # Set Tesseract path for Windows
        if os.name == 'nt':
            tesseract_paths = [
                r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"
            ]
            for path in tesseract_paths:
                if os.path.exists(path):
                    pytesseract.pytesseract.tesseract_cmd = path
                    break
        OCR_AVAILABLE = True
    except ImportError:
        OCR_AVAILABLE = False
except Exception:
    log_error("IMPORT ERROR:")
    log_error(traceback.format_exc())
    sys.exit(1)

app = FastAPI()

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# MongoDB Connection
MONGO_URI = "mongodb://localhost:27017/"
client = MongoClient(MONGO_URI)
db = client["rag_app_db"]

# Collections organized by document type for better performance
# Default collections
DEFAULT_COLLECTIONS = {
    "pdf": "documents_pdf",
    "text": "documents_text",
    "code": "documents_code",
    "office": "documents_office",
    "other": "documents_other"
}

# Initialize collections dictionary with defaults
collections = {name: db[coll_name] for name, coll_name in DEFAULT_COLLECTIONS.items()}

def get_all_collection_names():
    """Get all collection names from MongoDB"""
    return [name for name in db.list_collection_names() if name.startswith('documents_')]

def load_all_collections():
    """Load all collections including custom ones"""
    global collections
    all_coll_names = get_all_collection_names()
    collections = {}
    
    # Add default collections
    for name, coll_name in DEFAULT_COLLECTIONS.items():
        collections[name] = db[coll_name]
    
    # Add custom collections
    for coll_name in all_coll_names:
        if coll_name not in DEFAULT_COLLECTIONS.values():
            # Extract custom name (remove 'documents_' prefix)
            custom_name = coll_name.replace('documents_', '')
            if custom_name not in collections:
                collections[custom_name] = db[coll_name]
    
    return collections

# Load all collections on startup
load_all_collections()

def get_collection_for_file(filename: str):
    """Determine which collection to use based on file type"""
    file_lower = filename.lower()
    if file_lower.endswith(".pdf"):
        return collections["pdf"]
    elif file_lower.endswith((".txt", ".md", ".markdown")):
        return collections["text"]
    elif file_lower.endswith((".py", ".js", ".jsx", ".ts", ".tsx", ".java", ".cpp", ".c", ".h", ".cs", ".go", ".rs", ".rb", ".php", ".json", ".xml", ".yaml", ".yml", ".html", ".css")):
        return collections["code"]
    elif file_lower.endswith((".docx", ".doc", ".xlsx", ".xls")):
        return collections["office"]
    else:
        return collections["other"]

# Ollama Setup
OLLAMA_BASE_URL = "http://localhost:11434"

# Ensure upload directory exists in user's AppData to avoid permission issues
UPLOAD_DIR = os.path.join(os.getenv('APPDATA'), 'RAGulea', 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)



class QueryRequest(BaseModel):
    query: str
    model: str
    embedding_model: Optional[str] = "mxbai-embed-large:latest"
    collection_filter: Optional[List[str]] = None  # Filter by collection types

class ModelListResponse(BaseModel):
    models: List[str]

class CollectionStats(BaseModel):
    pdf: int
    text: int
    code: int
    office: int
    other: int
    total: int

class CreateCollectionRequest(BaseModel):
    name: str

def get_embeddings(text: str, model: str):
    embeddings = OllamaEmbeddings(model=model, base_url=OLLAMA_BASE_URL)
    return embeddings.embed_query(text)

@app.get("/api/models")
def get_models():
    # Fetch models from Ollama
    import requests
    try:
        response = requests.get(f"{OLLAMA_BASE_URL}/api/tags")
        if response.status_code == 200:
            models = [m["name"] for m in response.json()["models"]]
            return {"models": models}
    except Exception as e:
        return {"models": [], "error": str(e)}
    return {"models": []}

@app.get("/api/collections/stats")
def get_collection_stats():
    """Get document counts for each collection"""
    stats = {}
    total = 0
    for name, coll in collections.items():
        count = coll.count_documents({})
        stats[name] = count
        total += count
    stats["total"] = total
    return stats

@app.delete("/api/collections/{collection_name}")
def clear_collection(collection_name: str):
    """Clear all documents from a specific collection"""
    if collection_name not in collections:
        raise HTTPException(status_code=404, detail="Collection not found")
    result = collections[collection_name].delete_many({})
    return {"deleted": result.deleted_count}

@app.delete("/api/collections")
def clear_all_collections():
    """Clear all documents from all collections"""
    total_deleted = 0
    for coll in collections.values():
        result = coll.delete_many({})
        total_deleted += result.deleted_count
    return {"deleted": total_deleted}

@app.post("/api/collections/create")
def create_collection(request: CreateCollectionRequest):
    """Create a new custom collection"""
    print(f"📝 Received create collection request: {request}")
    name = request.name
    print(f"📝 Collection name: {name}")
    
    # Validate collection name
    if not name or not name.strip():
        print("❌ Collection name is empty")
        raise HTTPException(status_code=400, detail="Collection name cannot be empty")
    
    # Sanitize name (lowercase, alphanumeric + underscore only)
    sanitized_name = ''.join(c.lower() if c.isalnum() or c == '_' else '_' for c in name.strip())
    print(f"📝 Sanitized name: {sanitized_name}")
    
    if sanitized_name in collections:
        print(f"❌ Collection already exists: {sanitized_name}")
        raise HTTPException(status_code=400, detail="Collection already exists")
    
    # Create collection in MongoDB
    coll_name = f"documents_{sanitized_name}"
    collections[sanitized_name] = db[coll_name]
    print(f"✅ Created MongoDB collection: {coll_name}")
    
    # Create an index for better performance
    collections[sanitized_name].create_index("embedding_model")
    print(f"✅ Created index for collection: {sanitized_name}")
    
    return {
        "status": "success",
        "collection_name": sanitized_name,
        "mongodb_collection": coll_name
    }

@app.get("/api/collections/list")
def list_all_collections():
    """List all available collections including custom ones"""
    load_all_collections()  # Refresh collections
    
    collection_info = []
    for name, coll in collections.items():
        count = coll.count_documents({})
        is_default = name in DEFAULT_COLLECTIONS
        collection_info.append({
            "name": name,
            "count": count,
            "is_default": is_default,
            "mongodb_name": coll.name
        })
    
    return {"collections": collection_info}

@app.delete("/api/collections/custom/{collection_name}")
def delete_custom_collection(collection_name: str):
    """Delete a custom collection (cannot delete default collections)"""
    if collection_name in DEFAULT_COLLECTIONS:
        raise HTTPException(status_code=400, detail="Cannot delete default collections")
    
    if collection_name not in collections:
        raise HTTPException(status_code=404, detail="Collection not found")
    
    # Drop the collection from MongoDB
    coll_name = f"documents_{collection_name}"
    db.drop_collection(coll_name)
    
    # Remove from collections dict
    del collections[collection_name]
    
    return {"status": "success", "deleted": collection_name}

@app.post("/api/upload")
async def upload_file(
    file: UploadFile = File(...), 
    embedding_model: str = "mxbai-embed-large:latest",
    target_collection: Optional[str] = None
):
    file_path = os.path.join(UPLOAD_DIR, file.filename)
    
    print(f"\n📤 UPLOAD REQUEST:")
    print(f"   File: {file.filename}")
    print(f"   Embedding Model: {embedding_model}")
    print(f"   Target Collection: {target_collection}")
    
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        print(f"📁 Processing file: {file.filename}")
        
        # Process file based on extension
        text = ""
        file_lower = file.filename.lower()
        
        if file_lower.endswith(".pdf"):
            try:
                print(f"📄 Reading PDF: {file_path}")
                doc = fitz.open(file_path)
                page_count = len(doc)
                print(f"📄 PDF has {page_count} pages")
                text_parts = []
                
                # Try to extract text normally first
                for page_num in range(page_count):
                    page = doc[page_num]
                    page_text = page.get_text()
                    if page_text and page_text.strip():
                        text_parts.append(page_text)
                        print(f"   Page {page_num+1}: {len(page_text)} characters")
                
                doc.close()
                text = "\n".join(text_parts)
                print(f"📄 Total text extracted: {len(text)} characters")
                
                # If no text found, try OCR
                if not text.strip():
                    if not OCR_AVAILABLE:
                        raise HTTPException(status_code=400, detail="PDF contains scanned images. OCR libraries not installed. Run: pip install pytesseract pillow")
                    
                    print("📄 No text found, attempting OCR...")
                    try:
                        # Reopen PDF and extract images for OCR
                        doc = fitz.open(file_path)
                        ocr_text_parts = []
                        
                        for page_num in range(min(page_count, 50)):  # Limit to first 50 pages for performance
                            page = doc[page_num]
                            # Get page as image
                            pix = page.get_pixmap(dpi=200)
                            img_data = pix.tobytes("png")
                            
                            # Convert to PIL Image
                            from io import BytesIO
                            image = Image.open(BytesIO(img_data))
                            
                            # OCR the image
                            page_text = pytesseract.image_to_string(image, lang='ron+eng')  # Romanian + English
                            if page_text.strip():
                                ocr_text_parts.append(page_text)
                                print(f"   OCR Page {page_num+1}: {len(page_text)} characters")
                        
                        doc.close()
                        text = "\n".join(ocr_text_parts)
                        print(f"📄 Total OCR text extracted: {len(text)} characters")
                        
                        if not text.strip():
                            raise HTTPException(status_code=400, detail="PDF appears to be empty even after OCR")
                    except Exception as ocr_error:
                        print(f"❌ OCR Error: {str(ocr_error)}")
                        import traceback
                        traceback.print_exc()
                        raise HTTPException(status_code=400, detail=f"OCR failed: {str(ocr_error)}. Make sure Tesseract is installed: https://github.com/UB-Mannheim/tesseract/wiki")
                        
            except HTTPException:
                raise
            except Exception as e:
                print(f"❌ PDF Error: {str(e)}")
                print(f"❌ Error type: {type(e).__name__}")
                import traceback
                traceback.print_exc()
                raise HTTPException(status_code=400, detail=f"PDF processing failed: {str(e)}")
        
        elif file_lower.endswith((".txt", ".md", ".markdown")):
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    text = f.read()
            except UnicodeDecodeError:
                with open(file_path, "r", encoding="latin-1") as f:
                    text = f.read()
        
        elif file_lower.endswith((".py", ".js", ".jsx", ".ts", ".tsx", ".java", ".cpp", ".c", ".h", ".cs", ".go", ".rs", ".rb", ".php")):
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    text = f.read()
            except UnicodeDecodeError:
                with open(file_path, "r", encoding="latin-1") as f:
                    text = f.read()
        
        elif file_lower.endswith((".json", ".xml", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".conf")):
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    text = f.read()
            except UnicodeDecodeError:
                with open(file_path, "r", encoding="latin-1") as f:
                    text = f.read()
        
        elif file_lower.endswith((".html", ".htm", ".css", ".scss", ".sass")):
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    text = f.read()
            except UnicodeDecodeError:
                with open(file_path, "r", encoding="latin-1") as f:
                    text = f.read()
        
        elif file_lower.endswith((".csv", ".tsv")):
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    text = f.read()
            except UnicodeDecodeError:
                with open(file_path, "r", encoding="latin-1") as f:
                    text = f.read()
        
        elif file_lower.endswith((".docx", ".doc")):
            if not DOCX_AVAILABLE:
                raise HTTPException(status_code=400, detail="Word document support not installed. Run: pip install python-docx")
            try:
                doc = DocxDocument(file_path)
                text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Word document processing failed: {str(e)}")
        
        elif file_lower.endswith((".xlsx", ".xls")):
            if not EXCEL_AVAILABLE:
                raise HTTPException(status_code=400, detail="Excel support not installed. Run: pip install openpyxl")
            try:
                wb = load_workbook(file_path, data_only=True)
                text_parts = []
                for sheet in wb.worksheets:
                    text_parts.append(f"Sheet: {sheet.title}\n")
                    for row in sheet.iter_rows(values_only=True):
                        row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                        if row_text.strip():
                            text_parts.append(row_text)
                text = "\n".join(text_parts)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Excel processing failed: {str(e)}")
        
        else:
            # Try as text file
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    text = f.read()
            except:
                raise HTTPException(status_code=400, detail=f"Unsupported file type: {file.filename}")
        
        if not text or len(text.strip()) == 0:
            raise HTTPException(status_code=400, detail="File is empty or could not be read")
                
        # Split text
        text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
        chunks = text_splitter.split_text(text)
        
        if len(chunks) == 0:
            raise HTTPException(status_code=400, detail="No content to process after splitting")
        
        # Embed and store in appropriate collection
        embeddings_model = OllamaEmbeddings(model=embedding_model, base_url=OLLAMA_BASE_URL)
        
        # Use specified collection or auto-detect
        if target_collection and target_collection in collections:
            collection_to_use = collections[target_collection]
            print(f"✅ Using specified collection: {target_collection} ({collection_to_use.name})")
        else:
            collection_to_use = get_collection_for_file(file.filename)
            print(f"🔄 Auto-detected collection: {collection_to_use.name}")
            if target_collection:
                print(f"⚠️  Requested collection '{target_collection}' not found, using auto-detect")
        
        for chunk in chunks:
            vector = embeddings_model.embed_query(chunk)
            doc = {
                "filename": file.filename,
                "content": chunk,
                "embedding": vector,
                "embedding_model": embedding_model
            }
            collection_to_use.insert_one(doc)
            
        return {"status": "success", "chunks_processed": len(chunks)}
    
    except HTTPException:
        raise
    except Exception as e:
        log_error(f"Upload error for {file.filename}: {str(e)}")
        log_error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

def cosine_similarity(a, b):
    return np.dot(a, b) / (np.linalg.norm(a) * np.linalg.norm(b))

@app.post("/api/chat")
def chat(request: QueryRequest):
    print(f"\n🔍 CHAT REQUEST:")
    print(f"   Query: {request.query}")
    print(f"   Model: {request.model}")
    print(f"   Embedding Model: {request.embedding_model}")
    print(f"   Collection Filter: {request.collection_filter}")
    
    # Embed query
    embeddings_model = OllamaEmbeddings(model=request.embedding_model, base_url=OLLAMA_BASE_URL)
    query_vector = embeddings_model.embed_query(request.query)
    
    # Determine which collections to search
    collections_to_search = collections.values()
    if request.collection_filter:
        collections_to_search = [collections[name] for name in request.collection_filter if name in collections]
        print(f"   Searching collections: {request.collection_filter}")
    else:
        print(f"   Searching ALL collections ({len(collections)} total)")
    
    # Retrieve documents from relevant collections only
    results = []
    total_docs_searched = 0
    for coll in collections_to_search:
        cursor = coll.find({"embedding_model": request.embedding_model})
        coll_docs = 0
        for doc in cursor:
            coll_docs += 1
            if "embedding" in doc:
                score = cosine_similarity(query_vector, doc["embedding"])
                results.append((score, doc["content"], doc.get("filename", "unknown")))
        total_docs_searched += coll_docs
        if coll_docs > 0:
            print(f"   📁 {coll.name}: {coll_docs} documents")
    
    print(f"   Total documents searched: {total_docs_searched}")
    
    # Sort and get top results
    results.sort(key=lambda x: x[0], reverse=True)
    top_k = results[:5]
    
    print(f"   Top {len(top_k)} results:")
    for i, (score, content, filename) in enumerate(top_k):
        print(f"      {i+1}. Score: {score:.4f} | File: {filename} | Preview: {content[:100]}...")
    
    context = "\n\n".join([r[1] for r in top_k])
    
    # Generate response
    llm = OllamaLLM(model=request.model, base_url=OLLAMA_BASE_URL)
    prompt = f"Context:\n{context}\n\nQuestion: {request.query}\n\nAnswer:"
    response = llm.invoke(prompt)
    
    return {"response": response, "context": [r[1] for r in top_k]}

# Serve Frontend
# In development, we might run separately, but for the final app, we serve static files.
# We check if the dist folder exists relative to this file or the executable.
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

frontend_base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "frontend", "dist"))
print(f"Frontend base path: {frontend_base_path}")
print(f"Base path exists: {os.path.exists(frontend_base_path)}")

# When packaged with PyInstaller, the frontend files are in a `frontend` directory
# at the root of the temporary directory.
if getattr(sys, 'frozen', False):
    frontend_base_path = os.path.join(sys._MEIPASS, 'frontend')
    print(f"Packaged frontend base path: {frontend_base_path}")
    print(f"Packaged base path exists: {os.path.exists(frontend_base_path)}")

if os.path.exists(frontend_base_path):
    # Mount / to serve index.html
    @app.get("/")
    async def serve_index():
        return FileResponse(os.path.join(frontend_base_path, "index.html"))

    # Mount /vite.svg directly
    @app.get("/vite.svg")
    async def serve_vite_svg():
        return FileResponse(os.path.join(frontend_base_path, "vite.svg"))

    # Mount /assets for other static files
    app.mount("/assets", StaticFiles(directory=os.path.join(frontend_base_path, "assets")), name="static_assets")

if __name__ == "__main__":
    try:
        import socket as _socket
        import webbrowser
        import threading
        import requests
        
        print("=" * 60)
        print("RAGulea - Starting up...")
        print("=" * 60)
        
        # Check prerequisites
        print("\n🔍 Checking prerequisites...")
        
        # Check MongoDB
        try:
            client.server_info()
            print("✅ MongoDB: Connected")
        except Exception as e:
            print("❌ MongoDB: NOT RUNNING")
            print("\n⚠️  ERROR: MongoDB is not running!")
            print("   Please install and start MongoDB:")
            print("   Download: https://www.mongodb.com/try/download/community")
            print("\n   Or start MongoDB service:")
            print("   > net start MongoDB")
            input("\nPress Enter to exit...")
            sys.exit(1)
        
        # Check Ollama
        try:
            response = requests.get(f"{OLLAMA_BASE_URL}/api/tags", timeout=2)
            if response.status_code == 200:
                models = response.json().get("models", [])
                print(f"✅ Ollama: Connected ({len(models)} models available)")
                if len(models) == 0:
                    print("   ⚠️  Warning: No models found. Please pull models:")
                    print("   > ollama pull mxbai-embed-large")
                    print("   > ollama pull llama3")
            else:
                raise Exception("Ollama not responding")
        except Exception as e:
            print("❌ Ollama: NOT RUNNING")
            print("\n⚠️  ERROR: Ollama is not running!")
            print("   Please install Ollama:")
            print("   Download: https://ollama.ai/download")
            print("\n   Or start Ollama:")
            print("   > ollama serve")
            input("\nPress Enter to exit...")
            sys.exit(1)
        
        # Find available port
        chosen_port = None
        for p in range(8000, 8101):
            s = _socket.socket(_socket.AF_INET, _socket.SOCK_STREAM)
            try:
                s.bind(("0.0.0.0", p))
                s.close()
                chosen_port = p
                break
            except OSError:
                s.close()
                continue
        if chosen_port is None:
            chosen_port = 0
        
        print("\n" + "=" * 60)
        print(f"🚀 RAGulea Server Starting...")
        print(f"📍 Server: http://localhost:{chosen_port}")
        print(f"🎨 Frontend: {'AVAILABLE' if os.path.exists(frontend_base_path) else 'NOT FOUND'}")
        print("=" * 60)
        print("\n✨ Opening browser in 2 seconds...")
        print("   (You can close this window to stop the server)\n")
        
        # Open browser after a short delay to let server start
        def open_browser():
            import time
            time.sleep(2)
            webbrowser.open(f"http://localhost:{chosen_port}")
        
        threading.Thread(target=open_browser, daemon=True).start()
        
        uvicorn.run(app, host="0.0.0.0", port=chosen_port, log_config=None)
    except Exception:
        log_error("RUNTIME ERROR:")
        log_error(traceback.format_exc())
        print("An error occurred. Check ragulea_debug.log for details.")
        input("Press Enter to exit...")


